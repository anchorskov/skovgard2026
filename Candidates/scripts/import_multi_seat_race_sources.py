#!/usr/bin/env python3
"""Generate idempotent seed SQL for multi_seat_race_sources from the
multi-seat candidates flow spreadsheet.

Usage:
    python3 Candidates/scripts/import_multi_seat_race_sources.py > Candidates/db/seed/multi_seat_race_sources_import_YYYY-MM-DD.sql

Requires: 0019_multi_seat_race_sources.sql already applied (see docs/county_seed.md
style workflow — this project's migrations are applied by hand with
`wrangler d1 execute wy --remote --file=...`, not `wrangler d1 migrations apply`;
see the "shared wy D1" note in project memory for why).

Source: reads the 'multi_seat_import' and 'manual_review' sheets from
wy_2026_multi_seat_candidates_flow.xlsx in Downloads. Field names in that
sheet already follow this repo's naming standard (see
docs/data_import_field_standards.md) because the sheet was generated to spec.

Safe to re-run: every row is keyed by ballot_group_key and upserted with
INSERT ... ON CONFLICT DO UPDATE, so re-running against a refreshed
spreadsheet (new proclamations, corrected seat counts) updates existing rows
in place instead of duplicating them. Matching to live offices is a
separate, later step — see match_multi_seat_race_sources.py.
"""
import sys
from datetime import date

try:
    import openpyxl
except ImportError:
    sys.exit("Requires openpyxl: pip install openpyxl")

XLSX_PATH = '/mnt/c/Users/ancho/Downloads/wy_2026_multi_seat_candidates_flow.xlsx'
IMPORT_SHEET = 'multi_seat_import'
MANUAL_REVIEW_SHEET = 'manual_review'


def esc(v):
    if v is None:
        return 'NULL'
    s = str(v).strip()
    if not s:
        return 'NULL'
    return "'" + s.replace("'", "''") + "'"


def esc_int(v, default=0):
    if v is None or str(v).strip() == '':
        return str(default)
    return str(int(v))


def slugify(text):
    return ''.join(c if c.isalnum() else '_' for c in text.strip().lower()).strip('_')


def sheet_rows(wb, name):
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    idx = {h: i for i, h in enumerate(header)}
    return idx, rows[1:]


def build_import_rows(wb):
    idx, rows = sheet_rows(wb, IMPORT_SHEET)
    statements = []
    for r in rows:
        get = lambda field: r[idx[field]] if idx[field] < len(r) else None
        cols = {
            'ballot_group_key': get('ballot_group_key'),
            'county': get('county'),
            'election_phase': get('election_phase'),
            'jurisdiction_type': get('jurisdiction_type'),
            'city_or_town': get('city_or_town'),
            'precinct': get('precinct'),
            'precinct_name': get('precinct_name'),
            'party': get('party'),
            'office_name': get('office_name'),
            'district_or_scope': get('district_or_scope'),
            'term': get('term'),
            'seats_open': get('seats_open'),
            'max_selections': get('max_selections'),
            'ui_instruction': get('ui_instruction'),
            'source_type': get('source_type'),
            'source_url': get('source_url'),
            'source_status': get('source_status'),
            'notes': get('notes'),
        }
        if not cols['ballot_group_key']:
            continue
        statements.append(cols)
    return statements


def build_manual_review_rows(wb):
    if MANUAL_REVIEW_SHEET not in wb.sheetnames:
        return []
    idx, rows = sheet_rows(wb, MANUAL_REVIEW_SHEET)
    statements = []
    for r in rows:
        get = lambda field: r[idx[field]] if idx[field] < len(r) else None
        county = get('county')
        if not county:
            continue
        statements.append({
            'ballot_group_key': f"{slugify(county)}__manual_review",
            'county': county,
            'election_phase': None,
            'jurisdiction_type': 'Manual Review',
            'city_or_town': None,
            'precinct': None,
            'precinct_name': None,
            'party': None,
            'office_name': None,
            'district_or_scope': None,
            'term': None,
            'seats_open': 0,
            'max_selections': 0,
            'ui_instruction': None,
            'source_type': 'manual_review',
            'source_url': get('source_url'),
            'source_status': get('status'),
            'notes': get('notes'),
        })
    return statements


def emit_upsert(cols):
    fields = list(cols.keys())
    values = []
    for f in fields:
        if f in ('seats_open', 'max_selections'):
            values.append(esc_int(cols[f]))
        else:
            values.append(esc(cols[f]))
    update_clause = ', '.join(
        f"{f} = excluded.{f}" for f in fields if f != 'ballot_group_key'
    )
    return (
        f"INSERT INTO multi_seat_race_sources ({', '.join(fields)})\n"
        f"  VALUES ({', '.join(values)})\n"
        f"  ON CONFLICT(ballot_group_key) DO UPDATE SET\n"
        f"    {update_clause},\n"
        f"    updated_at = datetime('now');"
    )


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    import_rows = build_import_rows(wb)
    manual_rows = build_manual_review_rows(wb)

    print(f"-- Generated {date.today().isoformat()} by import_multi_seat_race_sources.py")
    print(f"-- Source: {XLSX_PATH.split('/')[-1]}")
    print(f"-- {len(import_rows)} race rows + {len(manual_rows)} manual-review placeholder rows")
    print("-- Safe to re-run: upserts by ballot_group_key.\n")

    for cols in import_rows + manual_rows:
        print(emit_upsert(cols))
        print()

    print(
        f"-- imported {len(import_rows)} race rows, {len(manual_rows)} manual-review rows",
        file=sys.stderr,
    )


if __name__ == '__main__':
    main()
