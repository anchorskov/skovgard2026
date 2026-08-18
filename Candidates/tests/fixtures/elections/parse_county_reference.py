# Candidates/tests/fixtures/elections/parse_county_reference.py
#
# Reference/scaffolding script only -- reproduces how every 2024-<county>/
# normalized/*.csv fixture was derived from the official 2024 XLSX workbooks
# (not checked in; re-download
# https://sos.wyo.gov/Elections/Docs/2024/Results/Primary/2024_Wyoming_Primary_Results.zip
# and place both workbooks next to this script to re-run). Not production
# parser code -- no error handling, no D1 writes, hardcoded to this one
# vendor's exact 3-row header layout. Treat it as a worked, validated example
# for what an `xlsx_wide_header` adapter needs to do -- see README.md.
import csv
import re
import sys
from collections import defaultdict
import openpyxl

SRC = "2024 Primary County PbP Results - OFFICIAL.xlsx"
BALLOTS_SRC = "2024 Primary County PbP Total Ballots Cast - OFFICIAL.xlsx"
ELECTION_DATE = "2024-08-20"
RETRIEVED_AT = "2026-08-18T18:00:00Z"


def office_level_and_district(contest_name):
    contest_name = re.sub(r",\s*Continued$", "", contest_name).strip()
    m = re.match(r"^(Senate|House) District (\d+)$", contest_name)
    if m:
        level = "wy_senate" if m.group(1) == "Senate" else "wy_house"
        return contest_name, level, int(m.group(2))
    if contest_name in ("United States Senator", "United States Representative"):
        return contest_name, "federal", None
    return contest_name, "unknown", None


def parse_county(county, pdf_source_url):
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    ws = wb[county]
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    contest_row, party_row, cand_row = rows[2], rows[3], rows[4]

    columns = []
    last_contest, last_party = None, None
    for i, (c, p, cand) in enumerate(zip(contest_row, party_row, cand_row)):
        if c:
            last_contest = c
        if p:
            last_party = p
        if i == 0 or cand is None:
            continue
        contest_name, level, district = office_level_and_district(last_contest)
        columns.append({
            "col": i, "contest": contest_name, "level": level, "district": district,
            "party": last_party,
            "candidate": cand.replace("\n", " ").strip() if isinstance(cand, str) else cand,
            "is_writein": cand == "Write-Ins", "is_overvote": cand == "Overvotes",
            "is_undervote": cand == "Undervotes",
        })

    def is_junk_label(v):
        # Page-break footer text ("Precincts Continue\non Next Page") leaks into
        # column A on sheets long enough to need an internal print break -- seen
        # on Natrona, Laramie, and Campbell but not the smaller Big Horn/Fremont
        # sheets. Never assume column A is a clean precinct code; filter by
        # what it explicitly isn't, not by position.
        return isinstance(v, str) and ("continue" in v.lower() or "page" in v.lower())

    data_rows = [r for r in rows[5:] if r[0] not in (None, "") and not is_junk_label(r[0])]
    total_row = next((r for r in data_rows if r[0] == "Total"), None)
    precinct_rows = [r for r in data_rows if r[0] != "Total"]
    precincts_total = len(precinct_rows)

    def row_type(col):
        return ("writein" if col["is_writein"] else "overvote" if col["is_overvote"]
                else "undervote" if col["is_undervote"] else "candidate")

    def build_rows(source_rows, precinct_field, source_url, source_format):
        out = []
        for prow in source_rows:
            precinct = prow[0] if precinct_field else ""
            for col in columns:
                val = prow[col["col"]]
                if val in ("-", None):
                    continue
                out.append({
                    "county": county, "election_date": ELECTION_DATE, "contest": col["contest"],
                    "office_level": col["level"], "district": col["district"] or "", "party": col["party"],
                    "precinct": precinct,
                    "candidate": col["candidate"] if row_type(col) == "candidate" else "",
                    "votes": val, "row_type": row_type(col),
                    "precincts_reporting": precincts_total, "precincts_total": precincts_total,
                    "is_unofficial": "false", "source_url": source_url, "source_format": source_format,
                    "retrieved_at_utc": RETRIEVED_AT,
                })
        return out

    precinct_out = build_rows(precinct_rows, True, pdf_source_url, "county_pbp_pdf")
    rollup_out = build_rows([total_row], False,
                             "https://sos.wyo.gov/Elections/Docs/2024/Results/Primary/2024_Wyoming_Primary_Results.zip",
                             "county_pbp_xlsx_official_zip")

    sums = defaultdict(int)
    for r in precinct_out:
        if r["row_type"] == "candidate":
            sums[(r["contest"], r["district"], r["candidate"])] += r["votes"]
    mismatches = []
    for r in rollup_out:
        if r["row_type"] == "candidate":
            key = (r["contest"], r["district"], r["candidate"])
            if sums.get(key) != r["votes"]:
                mismatches.append((key, sums.get(key), r["votes"]))

    # ballots cast sheet
    wb2 = openpyxl.load_workbook(BALLOTS_SRC, read_only=True, data_only=True)
    ws2 = wb2[county]
    brows = list(ws2.iter_rows(min_row=1, max_row=ws2.max_row, values_only=True))
    hidx = next(i for i, r in enumerate(brows) if r[0] == "Precinct")
    ballots = [r for r in brows[hidx + 1:] if r[0] not in (None, "") and r[0] != "Total" and not is_junk_label(r[0])]

    contests_seen = sorted(set((c["contest"], c["district"]) for c in columns if c["contest"]))

    return {
        "precincts_total": precincts_total,
        "precinct_rows": precinct_out,
        "rollup_rows": rollup_out,
        "mismatches": mismatches,
        "ballots": ballots,
        "contests": contests_seen,
    }


COUNTIES = {
    "Big Horn": ("bighorn", "https://sos.wyo.gov/Elections/Docs/2024/Results/Primary/2024_Big_Horn_County_Primary_PbP.pdf"),
    "Fremont": ("fremont", "https://sos.wyo.gov/Elections/Docs/2024/Results/Primary/2024_Fremont_County_Primary_PbP.pdf"),
    "Laramie": ("laramie", "https://sos.wyo.gov/Elections/Docs/2024/Results/Primary/2024_Laramie_County_Primary_PbP.pdf"),
    "Campbell": ("campbell", "https://sos.wyo.gov/Elections/Docs/2024/Results/Primary/2024_Campbell_County_Primary_PbP.pdf"),
    "Natrona": ("natrona", "https://sos.wyo.gov/Elections/Docs/2024/Results/Primary/2024_Natrona_County_Primary_PbP.pdf"),
}

if __name__ == "__main__":
    for county, (slug, pdf_url) in COUNTIES.items():
        result = parse_county(county, pdf_url)
        print(f"=== {county} ===")
        print(f"  precincts_total={result['precincts_total']}")
        print(f"  contests={len(result['contests'])}: {[c[0]+('#'+str(c[1]) if c[1] else '') for c in result['contests']]}")
        print(f"  precinct rows={len(result['precinct_rows'])}  rollup rows={len(result['rollup_rows'])}")
        print(f"  ballots-cast rows={len(result['ballots'])}")
        print(f"  reconciliation mismatches={len(result['mismatches'])}")
        for m in result["mismatches"][:5]:
            print("    MISMATCH", m)

        fieldnames = list(result["precinct_rows"][0].keys())
        with open(f"{slug}_2024_precinct_normalized.csv", "w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=fieldnames)
            w.writeheader(); w.writerows(result["precinct_rows"])
        with open(f"{slug}_2024_county_rollup.csv", "w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=fieldnames)
            w.writeheader(); w.writerows(result["rollup_rows"])
        with open(f"{slug}_2024_ballots_cast.csv", "w", newline="") as f:
            w = csv.writer(f)
            w.writerow(["precinct", "republican_ballots", "democratic_ballots", "nonpartisan_ballots", "total_ballots",
                        "county", "election_date", "source_url", "retrieved_at_utc"])
            for r in result["ballots"]:
                w.writerow([*r[:5], county, ELECTION_DATE,
                            "https://sos.wyo.gov/Elections/Docs/2024/Results/Primary/2024_Wyoming_Primary_Results.zip",
                            RETRIEVED_AT])

        # verbatim raw dump of the sheet (header rows + Total row, uncleaned)
        wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
        ws = wb[county]
        raw_rows = list(ws.iter_rows(min_row=1, max_row=5 + result["precincts_total"] + 1, values_only=True))
        with open(f"{slug}_raw_sheet_verbatim.csv", "w", newline="") as f:
            w = csv.writer(f)
            for r in raw_rows:
                w.writerow(["" if v is None else v for v in r])
        print()
